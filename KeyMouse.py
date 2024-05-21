import pyautogui  # 导入 pyautogui 库，用于模拟鼠标和键盘操作
import time  # 导入 time 库，用于控制时间延迟
import openpyxl  # 导入 openpyxl 库，用于读取和写入 Excel 文件
import pyperclip  # 导入 pyperclip 库，用于操作剪贴板
import os  # 导入 os 库，用于文件和操作系统的交互
import configparser  # 导入 configparser 库，用于读取配置文件

# 创建默认的 keyMouse.ini 配置文件
def create_default_ini():
    config = configparser.ConfigParser()
    config['Settings'] = {
        'timeDelay': '0.8',
        'executionMode': '1',
        'loopCount': '2',
        'excelFileName': 'orange.xlsx'  
    }
    with open('keyMouse.ini', 'w', encoding='utf-8') as configfile:
        config.write(configfile)

# 创建默认的 orange.xlsx 文件
def create_default_excel(excel_file_name):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['欢迎使用橙子草的Python自动化脚本~（https://github.com/Chinzicam/KeyMouse）'])
    sheet.append(['指令类型（1 单击  2 双击  3 右键  4 输入  5 等待  6滚轮 7 判断 8 键盘键入）',
                  '内容（图片名称.png、输入内容、等待时长/秒）', '重复次数(-1代表一直重复)'])
    
    # 创建数据行
    data_rows = [
        [8, 'win+r', 1],
        [4, 'cmd', 1],
        [8, 'enter', 1],
        [4, '欢迎使用橙子草的Python自动化脚本~', 1],
		[8, 'enter', 1],
		[4, '用法请查看 使用方法.txt~', 1]
    ]

    for row in data_rows:
        sheet.append(row)

    wb.save(excel_file_name)



# 检查并创建必要的文件和文件夹
if not os.path.exists('keyMouse.ini'):
    create_default_ini()

# 读取配置文件
config = configparser.ConfigParser()
config.read('keyMouse.ini', encoding='utf-8')

# 获取配置项
timeDelay = config.getfloat('Settings', 'timeDelay')
executionMode = config.getint('Settings', 'executionMode')
excelFileName = config.get('Settings', 'excelFileName') 

# 如果指定的 Excel 文件不存在，创建默认的 Excel 文件
if not os.path.exists(excelFileName):
    create_default_excel(excelFileName)

# 如果 img 文件夹不存在，创建该文件夹
if not os.path.exists('img'):
    os.makedirs('img')

# 生成 "使用方法.txt" 文件
def create_usage_file():
    usage_content = """欢迎使用橙子草的Python自动化脚本
使用方法
文件结构：
img文件夹：用于存放图片
cmd.xlsx：用于配置操作脚本
KeyMouse.py 或者 KeyMouse.exe：主文件

如何运行:双击启动exe/运行以下代码
python KeyMouse.py

其他
以下是一些功能8 常用键的名称，更多的键名称可以参考 pyautogui 的官方文档：
字母键：'a', 'b', 'c', ..., 'z'
数字键：'0', '1', '2', ..., '9'
功能键：'f1', 'f2', ..., 'f12'
箭头键：'left', 'right', 'up', 'down'
控制键：'ctrl', 'alt', 'shift', 'win', 'cmd' (Mac OS)
其他键：'enter', 'space', 'backspace', 'tab', 'esc', 'delete', 'home', 'end', 'pageup', 'pagedown'

项目地址：https://github.com/Chinzicam/KeyMouse，欢迎star~
"""
    with open('使用方法.txt', 'w', encoding='utf-8') as f:
        f.write(usage_content)

# 如果 "使用方法.txt" 文件不存在，创建该文件
if not os.path.exists('使用方法.txt'):
    create_usage_file()


# 如果是循环模式，读取循环次数
if executionMode == 2:
    loopCount = config.getint('Settings', 'loopCount')
else:
    loopCount = 1

# 定义鼠标事件函数
def mouseClick(clickTimes, lOrR, img, reTry):
    # 检查图像文件是否存在
    if not os.path.exists(img):
        print(f"文件 {img} 不存在，跳过该步骤")
        return
    
# 定义鼠标事件函数
def mouseClick(clickTimes, lOrR, img, reTry):
    # 检查图像文件是否存在
    if not os.path.exists(img):
        print(f"文件 {img} 不存在，跳过该步骤")
        return

    # 最多重试3次逻辑
    max_retries = 3
    retry_count = 0

    while retry_count < max_retries:
        location = pyautogui.locateCenterOnScreen(img, confidence=0.9)
        if location is not None:
            pyautogui.click(location.x, location.y, clicks=clickTimes, interval=0.2, duration=0.2, button=lOrR)
            break
        print(f"未找到匹配图片, 1秒后重试 (第{retry_count + 1}次)")
        retry_count += 1
        time.sleep(1)
    else:
        print(f"未找到匹配图片，超过最大重试次数{max_retries}，跳过该步骤")

    if reTry == -1:
        while True:
            location = pyautogui.locateCenterOnScreen(img, confidence=0.9)
            if location is not None:
                pyautogui.click(location.x, location.y, clicks=clickTimes, interval=0.2, duration=0.2, button=lOrR)
            time.sleep(0.1)
    elif reTry > 1:
        i = 1
        while i < reTry + 1:
            location = pyautogui.locateCenterOnScreen(img, confidence=0.9)
            if location is not None:
                pyautogui.click(location.x, location.y, clicks=clickTimes, interval=0.2, duration=0.2, button=lOrR)
                print("重复")
                i += 1
            time.sleep(0.1)


# 数据检查函数，确保 Excel 中的数据有效
def dataCheck(sheet):
    checkCmd = True
    if sheet.max_row < 3:
        print("没数据！")
        checkCmd = False
    for i in range(3, sheet.max_row + 1):
        cmdType = sheet.cell(row=i, column=1).value
        if cmdType not in [1, 2, 3, 4, 5, 6, 7, 8]:  # 确保操作类型在允许的范围内
            print(f'第{i}行,第1列数据有问题')
            checkCmd = False
        cmdValue = sheet.cell(row=i, column=2).value
        if cmdType in [1, 2, 3, 7, 8] and not isinstance(cmdValue, str):  # 确保特定类型的值是字符串
            print(f'第{i}行,第2列数据有问题')
            checkCmd = False
        if cmdType == 4 and not cmdValue:  # 确保类型4的值非空
            print(f'第{i}行,第2列数据有问题')
            checkCmd = False
        if cmdType in [5, 6] and not isinstance(cmdValue, (int, float)):  # 确保特定类型的值是数字
            print(f'第{i}行,第2列数据有问题')
            checkCmd = False
    return checkCmd

# 主函数，执行从 Excel 中读取的任务
def mainWork(sheet):
    for i in range(3, sheet.max_row + 1):
        cmdType = sheet.cell(row=i, column=1).value
        if cmdType == 1:
            img = 'img/' + sheet.cell(row=i, column=2).value
            reTry = sheet.cell(row=i, column=3).value or 1
            mouseClick(1, "left", img, int(reTry))
            print("单击左键", img)
        elif cmdType == 2:
            img = 'img/' + sheet.cell(row=i, column=2).value
            reTry = sheet.cell(row=i, column=3).value or 1
            mouseClick(2, "left", img, int(reTry))
            print("双击左键", img)
        elif cmdType == 3:
            img = 'img/' + sheet.cell(row=i, column=2).value
            reTry = sheet.cell(row=i, column=3).value or 1
            mouseClick(1, "right", img, int(reTry))
            print("右键", img)
        elif cmdType == 4:
            inputValue = sheet.cell(row=i, column=2).value
            pyperclip.copy(inputValue)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)
            print("输入:", inputValue)
        elif cmdType == 5:
            waitTime = sheet.cell(row=i, column=2).value
            time.sleep(waitTime)
            print("等待", waitTime, "秒")
        elif cmdType == 6:
            scroll = sheet.cell(row=i, column=2).value
            pyautogui.scroll(int(scroll))
            print("滚轮滑动", int(scroll), "距离")
        elif cmdType == 7:
            img = 'img/' + sheet.cell(row=i, column=2).value
            if not os.path.exists(img):
                print(f"文件 {img} 不存在，跳过该步骤")
                continue
            location = pyautogui.locateCenterOnScreen(img, confidence=0.9)
            if location is not None:
                print(f"图片 {img} 存在，执行后续操作")
                pyautogui.click(location.x, location.y)
            else:
                print(f"图片 {img} 不存在，跳过此步骤")
        elif cmdType == 8:
            keys = sheet.cell(row=i, column=2).value.split('+')
            pyautogui.hotkey(*keys)
            print("按键组合:", '+'.join(keys))
        time.sleep(timeDelay)  # 添加默认等待时间

if __name__ == '__main__':
    wb = openpyxl.load_workbook(filename=excelFileName)  # 加载 Excel 工作簿
    sheet = wb.active  # 获取活动工作表
    print('------------------------------~')
    print('欢迎使用橙子草的Python自动化脚本~')
    print('------------------------------~')
    checkCmd = dataCheck(sheet)  # 检查数据有效性
    if checkCmd:
        if executionMode == 1:
            mainWork(sheet)  # 执行一次
        elif executionMode == 2:
            for _ in range(loopCount):
                mainWork(sheet)  # 循环执行
                time.sleep(1)
                print("等待1秒")
    else:
        print('输入有误或者已经退出!')
